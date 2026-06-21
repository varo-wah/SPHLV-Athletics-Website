#!/usr/bin/env python3
"""Parse the SPHLV master sports schedule workbook into frontend JSON."""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


DEFAULT_INPUT = Path("/Users/williamhartono/Desktop/SPHLV Sports Master Schedule 2026-27.xlsx")
DEFAULT_OUTPUT = Path("src/data/schedule.json")
SEASONS = ("Season 1", "Season 2", "Season 3", "Season 4")
DAY_NAMES = ("Mon", "Tues", "Tue", "Wed", "Thur", "Thu", "Fri", "Sat", "Sun")
EVENT_TYPES = ("Practice", "Home Game", "Away Game", "Tournament", "Holiday", "Other")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(value))
    return None


def fill_rgb(cell: Cell) -> str:
    color = cell.fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb.upper()
    return ""


def is_week_label(value: Any) -> bool:
    return bool(re.search(r"\bWEEK\s*#?\s*\d+", clean(value), re.I))


def normalize_week(value: Any) -> str:
    text = clean(value).upper()
    match = re.search(r"WEEK\s*#?\s*(\d+)", text)
    return f"Week #{match.group(1)}" if match else text.title()


def is_day_label(value: Any) -> bool:
    text = clean(value).rstrip(".")
    return text in DAY_NAMES


def display_day(value: Any, parsed_date: date | None) -> str | None:
    if is_day_label(value):
        return clean(value).rstrip(".")
    if parsed_date:
        return parsed_date.strftime("%a")
    return None


def detect_event_type(text: str, fill: str, team: str) -> str:
    haystack = f"{text} {team}".lower()
    if fill == "FF7030A0" or any(
        token in haystack
        for token in ("holiday", "break", "no classes", "retreat", "pd day", "good friday", "academic holiday")
    ):
        return "Holiday"
    if any(token in haystack for token in ("tournament", "finals", "acsc", "jaac", "sph cup")):
        return "Tournament"
    if any(token in haystack for token in ("practice", "tryout", "pool recovery", "training", "meet in")):
        return "Practice"
    if re.search(r"\blv\s*@", haystack) or re.search(r"\bsph[-\s]?lv\s*@", haystack):
        return "Away Game"
    if "@ lv" in haystack or "@lv" in haystack or "sph-lv" in haystack or "sph lv" in haystack:
        return "Home Game"
    if "game" in haystack:
        return "Other"
    return "Other"


def detect_time(text: str, fallback: Any = None) -> str | None:
    from_text = re.search(r"\b\d{1,2}(?::\d{2})?\s*(?:am|pm)\b|\b\d{1,2}:\d{2}\b", text, re.I)
    if from_text:
        return from_text.group(0)
    fallback_text = clean(fallback)
    if fallback_text:
        return fallback_text
    return None


def detect_location(text: str) -> str | None:
    known = (
        "Main Field",
        "Side Field",
        "LV Field",
        "Gym 1",
        "Gym 2",
        "Covered Courts",
        "Covered",
        "Outdoor",
        "Track",
        "Pool",
        "K102",
        "SPH LV",
        "SPH-LV",
        "LV",
    )
    lower_text = text.lower()
    for venue in known:
        if venue.lower() in lower_text:
            return venue
    at_match = re.search(r"@\s*([A-Za-z0-9 -]+)", text)
    if at_match:
        return at_match.group(1).strip()
    return None


def detect_opponent(text: str) -> str | None:
    cleaned = clean(text)
    away_match = re.search(r"\bLV\s*@\s*([A-Za-z0-9 -]+)", cleaned, re.I)
    if away_match:
        return away_match.group(1).strip()
    home_match = re.search(r"([A-Za-z0-9 -]+)\s*@\s*LV\b", cleaned, re.I)
    if home_match:
        return home_match.group(1).strip()
    game_match = re.search(r"Game\s*-\s*([A-Za-z0-9 -]+?)(?:,|$)", cleaned, re.I)
    if game_match:
        return game_match.group(1).strip()
    return None


def header_map(ws) -> dict[int, str]:
    headers: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        value = clean(ws.cell(4, col).value)
        if not value:
            continue
        if value.upper().startswith("WEEK") or value in DAY_NAMES:
            continue
        headers[col] = value
    return headers


def parse_grid_season(ws, season: str) -> list[dict[str, Any]]:
    headers = header_map(ws)
    events: list[dict[str, Any]] = []
    current_week: str | None = None

    for row_idx in range(5, ws.max_row + 1):
        first = ws.cell(row_idx, 1).value
        second = ws.cell(row_idx, 2).value
        if is_week_label(first):
            current_week = normalize_week(first)
            continue

        parsed_date = date_value(second)
        day = display_day(first, parsed_date)
        if not parsed_date:
            continue

        for col_idx, team in headers.items():
            cell = ws.cell(row_idx, col_idx)
            text = clean(cell.value)
            if not text:
                continue

            event_type = detect_event_type(text, fill_rgb(cell), team)
            events.append(
                {
                    "id": f"{season.lower().replace(' ', '-')}-{row_idx}-{col_idx}",
                    "season": season,
                    "week": current_week,
                    "date": parsed_date.isoformat(),
                    "day": day,
                    "team": team,
                    "eventText": text,
                    "eventType": event_type,
                    "location": detect_location(text),
                    "opponent": detect_opponent(text),
                    "time": detect_time(text),
                    "raw": text,
                }
            )
    return events


def parse_season_four(ws, season: str) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    current_week: str | None = None
    current_date: date | None = None

    for row_idx in range(5, ws.max_row + 1):
        first = ws.cell(row_idx, 1).value
        if is_week_label(first):
            current_week = normalize_week(first)
            continue

        parsed_date = date_value(first) or current_date
        if date_value(first):
            current_date = date_value(first)
        if not parsed_date:
            continue

        fallback_time = ws.cell(row_idx, 2).value
        for col_idx in range(3, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            text = clean(cell.value)
            if not text:
                continue
            team = clean(ws.cell(4, col_idx).value) or "Season 4 Schedule"
            event_type = detect_event_type(text, fill_rgb(cell), team)
            events.append(
                {
                    "id": f"{season.lower().replace(' ', '-')}-{row_idx}-{col_idx}",
                    "season": season,
                    "week": current_week,
                    "date": parsed_date.isoformat(),
                    "day": parsed_date.strftime("%a"),
                    "team": team,
                    "eventText": text,
                    "eventType": event_type,
                    "location": detect_location(text),
                    "opponent": detect_opponent(text),
                    "time": detect_time(text, fallback_time),
                    "raw": text,
                }
            )
    return events


def main() -> int:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT

    wb = load_workbook(input_path, data_only=True)
    events: list[dict[str, Any]] = []
    warnings: list[str] = []

    for season in SEASONS:
        if season not in wb.sheetnames:
            warnings.append(f"Missing sheet: {season}")
            continue
        ws = wb[season]
        parsed = parse_season_four(ws, season) if season == "Season 4" else parse_grid_season(ws, season)
        if not parsed:
            warnings.append(f"No events parsed from {season}")
        events.extend(parsed)

    events.sort(key=lambda event: (event["date"] or "", event["season"], event["team"], event["eventText"]))
    payload = {
        "sourceFile": input_path.name,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "seasons": list(SEASONS),
        "eventTypes": list(EVENT_TYPES),
        "events": events,
        "warnings": warnings,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    for warning in warnings:
        print(f"WARNING: {warning}", file=sys.stderr)
    print(f"Wrote {len(events)} events to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
